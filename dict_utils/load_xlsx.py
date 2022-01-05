from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple


def xlsx_to_rows(file):
    wb = load_workbook(file)
    raw_rows = []
    for s_name in wb.sheetnames:
        sheet = wb[s_name]
        max_row, max_col = coordinate_to_tuple(sheet.dimensions.split(':')[1])
        for r in range(1, max_row + 1):
            row = []
            for col in range(1, max_col + 1):  # ignoring the first column containing the numbers
                value = sheet.cell(r, col).value
                row.append(value)
            raw_rows.append(row)
    return raw_rows


def rows_to_entries(rows):
    """
    supported format:
    - 1st row is legend
    - A column contains lemma
    - rows with empty A column are meanings of the last row that have a lemma in column A
    - other columns are optional definition fields
    """
    legend, rows = rows[0], rows[1:]
    entries = defaultdict(list)
    multi_entries = defaultdict(int)
    previous_lemma = ''
    for row in rows:
        acception = {legend[i]: row[i] for i in range(1, len(row)) if row[i]}

        lemma = row[0]
        if row[0] and lemma in entries:  # support for duplicate entries by adding a number at the end.
            multi_entries[lemma] += 1
            lemma = f'{lemma}_{multi_entries[lemma]}'
            previous_lemma = lemma
        elif not row[0] and previous_lemma:
            lemma = previous_lemma
        else:
            previous_lemma = lemma

        entries[lemma].append(acception)

    return entries


def load_xlsx_dict(file):
    rows = xlsx_to_rows(file)
    entries = rows_to_entries(rows)
    return entries
